import pandas
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from modules.paths import get_ind_nifty50list_csv_path, get_indice_csv_path, get_stock_csv_path, get_output_path
from modules.read_csv import read_symbols_from_csv, read_symbols_and_industries_from_csv, read_indice, \
    filter_dataframe_by_index_name, calculate_volume_increment, sort_dataframe, read_stock, filter_series_eq, \
    replace_series_based_on_symbol, calculate_volume_increment_stock

if __name__ == "__main__":

    ind_nifty50list_csv_path = get_ind_nifty50list_csv_path()

    user_date1 = input("User Date 1 as ddmmyyyy: ")
    user_date2 = input("User Date 2 as ddmmyyyy: ")

    # Nifty 50 industries set with stock name as key {'ADANIENT': 'Metals & Mining'}
    industries = read_symbols_and_industries_from_csv(ind_nifty50list_csv_path)

    if industries is not None:
        print(industries)

    indice_csv_path1 = get_indice_csv_path(user_date1)
    indice_csv_path2 = get_indice_csv_path(user_date2)
    indice_df1 = read_indice(indice_csv_path1)
    indice_df2 = read_indice(indice_csv_path2)
    filtered_df1 = filter_dataframe_by_index_name(indice_df1)
    filtered_df2 = filter_dataframe_by_index_name(indice_df2)

    # Assuming df1 and df2 are your two filtered dataframes
    vr = calculate_volume_increment(filtered_df1, filtered_df2)
    vr_sorted = sort_dataframe(vr, "Change(%)_df1")

    if indice_df1 is not None:
        print(vr_sorted)

    stock_csv_path1 = get_stock_csv_path(user_date1)
    stock_csv_path2 = get_stock_csv_path(user_date2)
    stock_df1 = read_stock(stock_csv_path1)
    stock_df2 = read_stock(stock_csv_path2)
    stock_eq1 = filter_series_eq(stock_df1, industries)
    stock_eq2 = filter_series_eq(stock_df2, industries)
    stock_industry1 = replace_series_based_on_symbol(stock_eq1,industries)
    stock_industry2 = replace_series_based_on_symbol(stock_eq2, industries)

    vr_stock = calculate_volume_increment_stock(stock_industry1,stock_industry2)
    vr_stock_sorted = sort_dataframe(vr_stock, 'Change % Previous')

    if stock_df1 is not None:
        print(vr_stock)

        output_file = f'{get_output_path()}/{user_date2}.xlsx'
        with pandas.ExcelWriter(output_file) as writer:
            # Write each dataframe to a different sheet
            vr_sorted.to_excel(writer, sheet_name='Indices', index=False)
            vr_stock_sorted.to_excel(writer, sheet_name='Nifty 50',index=False)

            workbook = writer.book
            worksheet = writer.sheets['Nifty 50']

            # Define styles
            # fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            font_top = Font(bold=True, size=14, color="FF35B04D")

            # Apply styles to the first 5 rows (excluding header)
            for row in range(2, 7):  # 2 to 6 are the first 5 rows in 1-indexed Excel
                for col in range(1, len(vr_stock_sorted.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    # cell.fill = fill
                    cell.font = font_top

            # Define styles for last 5 rows
            # red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            font_bottom = Font(bold=True, size=14, color="FFB03535")

            # Apply styles to the last 5 rows
            for row in range(len(vr_stock_sorted) + 2 - 5, len(vr_stock_sorted) + 2):  # last 5 rows in 1-indexed Excel
                for col in range(1, len(vr_stock_sorted.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    # cell.fill = red_fill
                    cell.font = font_bottom